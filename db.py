"""
QueryDeck — database layer.

Two engines are supported behind one interface:
  * "mysql"      -> pymysql (covers PlanetScale, MariaDB, and any MySQL that a
                    phpMyAdmin instance sits on top of)
  * "clickhouse" -> clickhouse-connect (ClickHouse Cloud / self-hosted)

Design goals that shape this file:
  * No certificate files, ever. TLS "just works" via the certifi CA bundle,
    with a graceful fall back to a plain connection for local servers that
    don't speak TLS. A per-connection "skip verification" switch handles
    self-signed certs.
  * Survive huge result sets. Previews are paginated (LIMIT/OFFSET). Downloads
    are streamed row-block by row-block so memory stays flat whether the query
    returns a thousand rows or a trillion.
"""

from __future__ import annotations

import csv
import datetime as _dt
import decimal
import io
import json
import os
import re
import threading
import time
import uuid
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator

import certifi
import clickhouse_connect
from dotenv import load_dotenv
import pymysql
import pymysql.cursors

# Load environment variables from .env file
load_dotenv()

# --------------------------------------------------------------------------- #
# Storage
# --------------------------------------------------------------------------- #

_HERE = Path(__file__).resolve().parent
_DATA_DIR = os.environ.get("QUERYDECK_DATA_DIR")
if _DATA_DIR:
    _CUSTOM_DIR = Path(_DATA_DIR)
    _CUSTOM_DIR.mkdir(parents=True, exist_ok=True)
    _CUSTOM_FILE = _CUSTOM_DIR / "connections.json"
else:
    _CUSTOM_FILE = _HERE / "connections.json"
_FILE_LOCK = threading.Lock()

# Fields that must never be shown back to the browser.
_SECRET_FIELDS = {"password"}

# Rows fetched per network round-trip while streaming a download.
_CHUNK = 5_000
# Excel hard limit is 1,048,576 rows incl. the header row.
_XLSX_MAX_ROWS = 1_048_575
# Preview safety ceiling so a fat page can't blow up the browser.
_PAGE_SIZE_CAP = 2_000


# --------------------------------------------------------------------------- #
# Seed connections (the four you provided). These always exist and can't be
# deleted from the UI; treat them as read-only presets. Move the secrets to
# environment variables before sharing this project — see the README.
# --------------------------------------------------------------------------- #

DEFAULT_CONNECTIONS: list[dict[str, Any]] = [
    {
        "id": "crm-db",
        "name": "CRM DB",
        "engine": "mysql",
        "host": "aws.connect.psdb.cloud",
        "port": 3306,
        "user": "aav384pz1rttlp43moxa",
        "password": os.environ.get("CRM_DB_PASSWORD", ""),
        "database": "crm-db",
        "provider": "PlanetScale",
        "secure": True,
        "skip_verify": False,
        "builtin": True,
    },
    {
        "id": "notification-shortlink-db",
        "name": "Notification / Shortlink",
        "engine": "mysql",
        "host": "aws.connect.psdb.cloud",
        "port": 3306,
        "user": "cjd2hjs7lbxfrmi7sgmn",
        "password": os.environ.get("NOTIFICATION_DB_PASSWORD", ""),
        "database": "notification-shortlink-db",
        "provider": "PlanetScale",
        "secure": True,
        "skip_verify": False,
        "builtin": True,
    },
    {
        "id": "clirnetdb",
        "name": "CLIRNET DB",
        "engine": "mysql",
        "host": "aws.connect.psdb.cloud",
        "port": 3306,
        "user": "1j8lruwaxnjf1h38w1qx",
        "password": os.environ.get("CLIRNET_DB_PASSWORD", ""),
        "database": "clirnetdb",
        "provider": "PlanetScale",
        "secure": True,
        "skip_verify": False,
        "builtin": True,
    },
    {
        "id": "clickhouse-analytics",
        "name": "ClickHouse Analytics",
        "engine": "clickhouse",
        "host": "ayz7jk0o0v.ap-south-1.aws.clickhouse.cloud",
        "port": 8443,
        "user": "mr_nirmalendu",
        "password": os.environ.get("CLICKHOUSE_DB_PASSWORD", ""),
        "database": "default",
        "provider": "ClickHouse Cloud",
        "secure": True,
        "skip_verify": False,
        "builtin": True,
    },
]


def _load_custom() -> list[dict[str, Any]]:
    if not _CUSTOM_FILE.exists():
        return []
    try:
        data = json.loads(_CUSTOM_FILE.read_text("utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_custom(items: list[dict[str, Any]]) -> None:
    with _FILE_LOCK:
        _CUSTOM_FILE.write_text(json.dumps(items, indent=2), "utf-8")


def all_connections() -> list[dict[str, Any]]:
    """Built-in presets first, then anything the user added."""
    return [dict(c) for c in DEFAULT_CONNECTIONS] + _load_custom()


def public_connections() -> list[dict[str, Any]]:
    """Same list, with secrets stripped — safe to send to the browser."""
    out = []
    for c in all_connections():
        pub = {k: v for k, v in c.items() if k not in _SECRET_FIELDS}
        pub["has_password"] = bool(c.get("password"))
        out.append(pub)
    return out


def get_connection(conn_id: str) -> dict[str, Any] | None:
    for c in all_connections():
        if c["id"] == conn_id:
            return c
    return None


class ConnectionError_(Exception):
    """Raised when a connection id is unknown."""


def require_connection(conn_id: str) -> dict[str, Any]:
    cfg = get_connection(conn_id)
    if cfg is None:
        raise ConnectionError_(f"Unknown connection: {conn_id!r}")
    return cfg


def _normalise_incoming(payload: dict[str, Any]) -> dict[str, Any]:
    engine = (payload.get("engine") or "mysql").strip().lower()
    if engine not in ("mysql", "clickhouse"):
        raise ValueError("engine must be 'mysql' or 'clickhouse'")

    secure = bool(payload.get("secure", True))
    default_port = (8443 if secure else 8123) if engine == "clickhouse" else 3306
    try:
        port = int(payload.get("port") or default_port)
    except (TypeError, ValueError):
        port = default_port

    host = (payload.get("host") or "").strip()
    if not host:
        raise ValueError("host is required")

    # Check if host is a Google Cloud SQL Instance Connection Name (format: project:region:instance)
    if len(host.split(":")) == 3:
        raise ValueError(
            "It looks like you entered a Google Cloud SQL Instance Connection Name. "
            "Standard drivers cannot connect directly to this. Please use the public/private "
            "IP of your database instance as the 'Host', or run the Cloud SQL Auth Proxy "
            "locally and set the 'Host' to '127.0.0.1' and 'Port' to the proxy port."
        )

    user = (payload.get("user") or "").strip()
    if not user and engine == "mysql":
        raise ValueError("user is required")

    return {
        "engine": engine,
        "name": (payload.get("name") or host).strip(),
        "host": host,
        "port": port,
        "user": user,
        "password": payload.get("password") or "",
        "database": (payload.get("database") or "").strip(),
        "provider": (payload.get("provider") or "").strip(),
        "secure": secure,
        "skip_verify": bool(payload.get("skip_verify", False)),
    }


def add_connection(payload: dict[str, Any]) -> dict[str, Any]:
    cfg = _normalise_incoming(payload)
    cfg["id"] = "conn_" + uuid.uuid4().hex[:10]
    cfg["builtin"] = False
    items = _load_custom()
    items.append(cfg)
    _save_custom(items)
    pub = {k: v for k, v in cfg.items() if k not in _SECRET_FIELDS}
    pub["has_password"] = bool(cfg.get("password"))
    return pub


def delete_connection(conn_id: str) -> None:
    if any(c["id"] == conn_id for c in DEFAULT_CONNECTIONS):
        raise PermissionError("Built-in connections can't be deleted.")
    items = _load_custom()
    kept = [c for c in items if c["id"] != conn_id]
    if len(kept) == len(items):
        raise ConnectionError_(f"Unknown connection: {conn_id!r}")
    _save_custom(kept)


# --------------------------------------------------------------------------- #
# Value coercion — turn driver-native values into JSON / CSV / XLSX-safe ones.
# --------------------------------------------------------------------------- #

def _json_cell(v: Any) -> Any:
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, decimal.Decimal):
        return str(v)                       # exact — no float rounding
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat(sep=" ") if isinstance(v, _dt.datetime) else v.isoformat()
    if isinstance(v, (bytes, bytearray)):
        try:
            return v.decode("utf-8")
        except Exception:
            return v.hex()
    return str(v)


def _csv_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (bytes, bytearray)):
        try:
            return v.decode("utf-8")
        except Exception:
            return v.hex()
    if isinstance(v, _dt.datetime):
        return v.isoformat(sep=" ")
    return v


def _xlsx_cell(v: Any) -> Any:
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, decimal.Decimal):
        return float(v)
    if isinstance(v, _dt.datetime):
        return v.replace(tzinfo=None)        # Excel can't store tz-aware datetimes
    if isinstance(v, (_dt.date, _dt.time)):
        return v
    if isinstance(v, (bytes, bytearray)):
        try:
            return v.decode("utf-8")
        except Exception:
            return v.hex()
    return str(v)


# --------------------------------------------------------------------------- #
# SQL helpers
# --------------------------------------------------------------------------- #

def _clean_sql_and_check_multi(sql: str) -> tuple[str, bool, str]:
    sql = (sql or "").strip()
    if not sql:
        return "", False, ""

    in_single_quote = False
    in_double_quote = False
    in_backtick = False
    in_line_comment = False
    in_block_comment = False

    last_semantic_idx = -1

    i = 0
    n = len(sql)
    out = []

    while i < n:
        c = sql[i]
        next_c = sql[i+1] if i + 1 < n else ''

        if in_line_comment:
            if c == '\n':
                in_line_comment = False
                out.append(c)
            i += 1
            continue

        if in_block_comment:
            if c == '*' and next_c == '/':
                in_block_comment = False
                i += 2
            else:
                i += 1
            continue

        if in_single_quote:
            if c == '\\':
                out.append(c)
                if next_c:
                    out.append(next_c)
                    last_semantic_idx = i + 1
                    i += 2
                    continue
            elif c == "'":
                in_single_quote = False
            out.append(c)
            last_semantic_idx = i
            i += 1
            continue

        if in_double_quote:
            if c == '\\':
                out.append(c)
                if next_c:
                    out.append(next_c)
                    last_semantic_idx = i + 1
                    i += 2
                    continue
            elif c == '"':
                in_double_quote = False
            out.append(c)
            last_semantic_idx = i
            i += 1
            continue

        if in_backtick:
            if c == '`':
                in_backtick = False
            out.append(c)
            last_semantic_idx = i
            i += 1
            continue

        if c == '-' and next_c == '-':
            in_line_comment = True
            i += 2
            continue
        if c == '#':
            in_line_comment = True
            i += 1
            continue
        if c == '/' and next_c == '*':
            in_block_comment = True
            i += 2
            continue

        if c == "'":
            in_single_quote = True
            last_semantic_idx = i
        elif c == '"':
            in_double_quote = True
            last_semantic_idx = i
        elif c == '`':
            in_backtick = True
            last_semantic_idx = i
        elif c not in (';', ' ', '\t', '\r', '\n'):
            last_semantic_idx = i

        out.append(c)
        i += 1

    cleaned = "".join(out).strip()
    while cleaned.endswith(';'):
        cleaned = cleaned[:-1].strip()

    has_middle_semicolon = False
    in_single_quote = False
    in_double_quote = False
    in_backtick = False
    i = 0
    while i < len(cleaned):
        c = cleaned[i]
        next_c = cleaned[i+1] if i + 1 < len(cleaned) else ''
        if in_single_quote:
            if c == '\\' and next_c:
                i += 2
            elif c == "'":
                in_single_quote = False
                i += 1
            else:
                i += 1
        elif in_double_quote:
            if c == '\\' and next_c:
                i += 2
            elif c == '"':
                in_double_quote = False
                i += 1
            else:
                i += 1
        elif in_backtick:
            if c == '`':
                in_backtick = False
            i += 1
        else:
            if c == "'":
                in_single_quote = True
            elif c == '"':
                in_double_quote = True
            elif c == '`':
                in_backtick = True
            elif c == ';':
                has_middle_semicolon = True
                break
            i += 1

    sliced_original = sql[:last_semantic_idx + 1] if last_semantic_idx != -1 else sql
    return cleaned, has_middle_semicolon, sliced_original


def _clean_sql(sql: str) -> str:
    _, _, sliced = _clean_sql_and_check_multi(sql)
    return sliced


def _is_wrappable(sql: str) -> bool:
    """A single SELECT/WITH statement can be wrapped for LIMIT/OFFSET paging."""
    cleaned, has_multi, _ = _clean_sql_and_check_multi(sql)
    if has_multi:
        return False
    return bool(re.match(r"(?is)^\s*(select|with)\b", cleaned))


# --------------------------------------------------------------------------- #
# MySQL driver
# --------------------------------------------------------------------------- #

def _mysql_connect(cfg: dict[str, Any], read_timeout: int | None):
    base = dict(
        host=cfg["host"],
        port=int(cfg.get("port") or 3306),
        user=cfg["user"],
        password=cfg.get("password") or "",
        database=cfg.get("database") or None,
        charset="utf8mb4",
        autocommit=True,
        connect_timeout=15,
        read_timeout=read_timeout,
        write_timeout=30,
        local_infile=False,
    )
    secure = cfg.get("secure", True)
    if not secure:
        conn = pymysql.connect(ssl_disabled=True, **base)
    else:
        ssl_opts: dict[str, Any] = {"ca": certifi.where()}
        kwargs = dict(base)
        kwargs["ssl"] = ssl_opts
        if cfg.get("skip_verify"):
            kwargs["ssl_verify_cert"] = False
            kwargs["ssl_verify_identity"] = False
        try:
            conn = pymysql.connect(**kwargs)
        except Exception:
            # Server may not support TLS at all (typical local MySQL). Fall back.
            conn = pymysql.connect(ssl_disabled=True, **base)

    try:
        with conn.cursor() as cur:
            cur.execute("SET workload = 'olap'")
    except Exception:
        pass

    return conn


def _mysql_schema(cfg: dict[str, Any]) -> dict[str, list[str]]:
    conn = _mysql_connect(cfg, read_timeout=60)
    try:
        cur = conn.cursor()
        db = cfg.get("database")
        if db:
            cur.execute(
                "SELECT TABLE_NAME, COLUMN_NAME FROM information_schema.COLUMNS "
                "WHERE TABLE_SCHEMA=%s ORDER BY TABLE_NAME, ORDINAL_POSITION",
                (db,),
            )
        else:
            cur.execute(
                "SELECT TABLE_NAME, COLUMN_NAME FROM information_schema.COLUMNS "
                "WHERE TABLE_SCHEMA=DATABASE() ORDER BY TABLE_NAME, ORDINAL_POSITION"
            )
        tables: dict[str, list[str]] = {}
        for t, c in cur.fetchall():
            tables.setdefault(t, []).append(c)
        return tables
    finally:
        conn.close()


def _mysql_query(cfg: dict[str, Any], sql: str, page: int, page_size: int) -> dict[str, Any]:
    sql = _clean_sql(sql)
    conn = _mysql_connect(cfg, read_timeout=120)
    started = time.perf_counter()
    try:
        cur = conn.cursor()
        if _is_wrappable(sql):
            offset = page * page_size
            cur.execute(
                f"SELECT * FROM (\n{sql}\n) AS _qd LIMIT {page_size + 1} OFFSET {offset}"
            )
            rows = cur.fetchall()
            has_next = len(rows) > page_size
            rows = rows[:page_size]
            cols = [d[0] for d in cur.description]
            return _rowset(cols, rows, has_next, page, page_size, started)

        cur.execute(sql)
        if cur.description:                                   # SHOW / DESCRIBE / EXPLAIN
            rows = cur.fetchmany(page_size)
            cols = [d[0] for d in cur.description]
            res = _rowset(cols, rows, False, 0, page_size, started)
            res["note"] = "Preview only — this statement can't be paged."
            return res

        affected = cur.rowcount                               # INSERT / UPDATE / DDL
        return {
            "columns": [], "rows": [], "has_next": False, "page": 0,
            "page_size": page_size, "elapsed_ms": _ms(started),
            "affected_rows": affected,
            "message": f"Statement executed — {affected} row(s) affected.",
        }
    finally:
        conn.close()


def _mysql_export_csv(cfg: dict[str, Any], sql: str) -> Iterator[bytes]:
    """Stream the full result as CSV using a server-side cursor."""
    sql = _clean_sql(sql)
    conn = _mysql_connect(cfg, read_timeout=None)
    cur = conn.cursor(pymysql.cursors.SSCursor)
    cur.execute(sql)
    cols = [d[0] for d in cur.description] if cur.description else []

    def gen() -> Iterator[bytes]:
        try:
            buf = io.StringIO()
            writer = csv.writer(buf)
            yield b"\xef\xbb\xbf"                              # UTF-8 BOM for Excel
            if cols:
                writer.writerow(cols)
                yield buf.getvalue().encode("utf-8")
                buf.seek(0); buf.truncate(0)
            while True:
                rows = cur.fetchmany(_CHUNK)
                if not rows:
                    break
                for r in rows:
                    writer.writerow([_csv_cell(v) for v in r])
                yield buf.getvalue().encode("utf-8")
                buf.seek(0); buf.truncate(0)
        finally:
            cur.close()
            conn.close()

    return gen()


def _mysql_export_xlsx(cfg: dict[str, Any], sql: str, path: str) -> dict[str, Any]:
    """Write the full result to an Excel file using a server-side cursor."""
    from openpyxl import Workbook

    sql = _clean_sql(sql)
    conn = _mysql_connect(cfg, read_timeout=None)
    cur = conn.cursor(pymysql.cursors.SSCursor)
    cur.execute(sql)
    cols = [d[0] for d in cur.description] if cur.description else []
    try:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Result")
        if cols:
            ws.append(cols)
        written = 0
        truncated = False
        while True:
            rows = cur.fetchmany(_CHUNK)
            if not rows:
                break
            for r in rows:
                if written >= _XLSX_MAX_ROWS:
                    truncated = True
                    break
                ws.append([_xlsx_cell(v) for v in r])
                written += 1
            if truncated:
                break
        if truncated:
            note = wb.create_sheet("Truncated")
            note.append(["This export hit Excel's 1,048,575-row limit."])
            note.append(["Use CSV export for the complete result set."])
        wb.save(path)
        return {"rows": written, "truncated": truncated}
    finally:
        cur.close()
        conn.close()


def _mysql_test(cfg: dict[str, Any]) -> dict[str, Any]:
    conn = _mysql_connect(cfg, read_timeout=20)
    try:
        cur = conn.cursor()
        cur.execute("SELECT VERSION()")
        version = cur.fetchone()[0]
        table_count = None
        if cfg.get("database"):
            cur.execute(
                "SELECT COUNT(*) FROM information_schema.TABLES WHERE TABLE_SCHEMA=%s",
                (cfg["database"],),
            )
            table_count = cur.fetchone()[0]
        detail = f"MySQL {version}"
        if table_count is not None:
            detail += f" · {table_count} table(s)"
        return {"ok": True, "message": "Connected", "detail": detail}
    finally:
        conn.close()


# --------------------------------------------------------------------------- #
# ClickHouse driver
# --------------------------------------------------------------------------- #

def _ch_client(cfg: dict[str, Any], timeout: int = 300):
    secure = cfg.get("secure", True)
    skip = cfg.get("skip_verify", False)
    return clickhouse_connect.get_client(
        host=cfg["host"],
        port=int(cfg.get("port") or (8443 if secure else 8123)),
        username=cfg["user"],
        password=cfg.get("password") or "",
        database=cfg.get("database") or "default",
        secure=secure,
        verify=(not skip),
        ca_cert=(certifi.where() if secure and not skip else None),
        connect_timeout=15,
        send_receive_timeout=timeout,
        query_limit=0,                        # 0 = no implicit row cap
        compress=True,
    )


def _ch_schema(cfg: dict[str, Any]) -> dict[str, list[str]]:
    client = _ch_client(cfg, timeout=60)
    try:
        db = cfg.get("database") or "default"
        res = client.query(
            "SELECT table, name FROM system.columns "
            "WHERE database = {db:String} ORDER BY table, position",
            parameters={"db": db},
        )
        tables: dict[str, list[str]] = {}
        for t, c in res.result_rows:
            tables.setdefault(t, []).append(c)
        return tables
    finally:
        client.close()


def _ch_query(cfg: dict[str, Any], sql: str, page: int, page_size: int) -> dict[str, Any]:
    sql = _clean_sql(sql)
    client = _ch_client(cfg, timeout=120)
    started = time.perf_counter()
    try:
        if _is_wrappable(sql):
            offset = page * page_size
            res = client.query(f"SELECT * FROM (\n{sql}\n) LIMIT {page_size + 1} OFFSET {offset}")
            rows = list(res.result_rows)
            has_next = len(rows) > page_size
            rows = rows[:page_size]
            return _rowset(res.column_names, rows, has_next, page, page_size, started)

        # DDL / DML / SHOW / DESCRIBE
        try:
            res = client.query(sql)
            if res.column_names:
                rows = list(res.result_rows)[:page_size]
                out = _rowset(res.column_names, rows, False, 0, page_size, started)
                out["note"] = "Preview only — this statement can't be paged."
                return out
        except Exception:
            pass
        summary = client.command(sql)
        return {
            "columns": [], "rows": [], "has_next": False, "page": 0,
            "page_size": page_size, "elapsed_ms": _ms(started),
            "message": f"Statement executed. {summary}".strip(),
        }
    finally:
        client.close()


def _ch_columns(client, sql: str) -> list[str]:
    """Validate a SELECT and grab its column names without pulling rows."""
    res = client.query(f"SELECT * FROM (\n{sql}\n) LIMIT 0")
    return list(res.column_names)


def _ch_export_csv(cfg: dict[str, Any], sql: str) -> Iterator[bytes]:
    sql = _clean_sql(sql)
    client = _ch_client(cfg, timeout=3600)
    try:
        if _is_wrappable(sql):
            _ch_columns(client, sql)                                  # eager validation
            query_sql = f"SELECT * FROM (\n{sql}\n)"
        else:
            query_sql = sql
        raw = client.raw_stream(query_sql, fmt="CSVWithNames")
    except Exception as e:
        client.close()
        raise e

    def gen() -> Iterator[bytes]:
        try:
            yield b"\xef\xbb\xbf"
            source = raw
            if hasattr(raw, "__enter__"):
                with raw as stream:
                    for chunk in stream:
                        if chunk:
                            yield chunk
            else:
                for chunk in source:
                    if chunk:
                        yield chunk
        finally:
            try:
                client.close()
            except Exception:
                pass

    return gen()


def _ch_export_xlsx(cfg: dict[str, Any], sql: str, path: str) -> dict[str, Any]:
    from openpyxl import Workbook

    sql = _clean_sql(sql)
    client = _ch_client(cfg, timeout=3600)
    try:
        if _is_wrappable(sql):
            cols = _ch_columns(client, sql)
            query_sql = f"SELECT * FROM (\n{sql}\n)"
            is_select = True
        else:
            res = client.query(sql)
            cols = list(res.column_names)
            is_select = False

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Result")
        if cols:
            ws.append(cols)
        written = 0
        truncated = False

        if is_select:
            with client.query_row_block_stream(query_sql) as stream:
                for block in stream:
                    for row in block:
                        if written >= _XLSX_MAX_ROWS:
                            truncated = True
                            break
                        ws.append([_xlsx_cell(v) for v in row])
                        written += 1
                    if truncated:
                        break
        else:
            for row in res.result_rows:
                if written >= _XLSX_MAX_ROWS:
                    truncated = True
                    break
                ws.append([_xlsx_cell(v) for v in row])
                written += 1

        if truncated:
            note = wb.create_sheet("Truncated")
            note.append(["This export hit Excel's 1,048,575-row limit."])
            note.append(["Use CSV export for the complete result set."])
        wb.save(path)
        return {"rows": written, "truncated": truncated}
    finally:
        client.close()


def _ch_test(cfg: dict[str, Any]) -> dict[str, Any]:
    client = _ch_client(cfg, timeout=20)
    try:
        version = client.command("SELECT version()")
        db = cfg.get("database") or "default"
        count = client.command(
            "SELECT count() FROM system.tables WHERE database = {db:String}",
            parameters={"db": db},
        )
        return {"ok": True, "message": "Connected",
                "detail": f"ClickHouse {version} · {count} table(s) in {db}"}
    finally:
        client.close()


# --------------------------------------------------------------------------- #
# Shared helpers + public dispatch
# --------------------------------------------------------------------------- #

def _ms(started: float) -> int:
    return int((time.perf_counter() - started) * 1000)


def _rowset(cols, rows, has_next, page, page_size, started) -> dict[str, Any]:
    return {
        "columns": list(cols),
        "rows": [[_json_cell(v) for v in row] for row in rows],
        "has_next": has_next,
        "page": page,
        "page_size": page_size,
        "elapsed_ms": _ms(started),
    }


def _clamp_page_size(n: int) -> int:
    try:
        n = int(n)
    except (TypeError, ValueError):
        n = 200
    return max(1, min(n, _PAGE_SIZE_CAP))


def run_query(cfg: dict[str, Any], sql: str, page: int = 0, page_size: int = 200) -> dict[str, Any]:
    page = max(0, int(page or 0))
    page_size = _clamp_page_size(page_size)
    if cfg["engine"] == "mysql":
        return _mysql_query(cfg, sql, page, page_size)
    return _ch_query(cfg, sql, page, page_size)


def get_schema(cfg: dict[str, Any]) -> dict[str, Any]:
    tables = _mysql_schema(cfg) if cfg["engine"] == "mysql" else _ch_schema(cfg)
    ordered = dict(sorted(tables.items()))
    return {
        "database": cfg.get("database") or "",
        "engine": cfg["engine"],
        "tables": [{"name": t, "columns": ordered[t]} for t in ordered],
        "hints": ordered,           # {table: [columns]} for the editor autocomplete
    }


def test_connection(payload: dict[str, Any]) -> dict[str, Any]:
    try:
        cfg = _normalise_incoming(payload)
    except ValueError as e:
        return {"ok": False, "message": str(e)}
    try:
        return _mysql_test(cfg) if cfg["engine"] == "mysql" else _ch_test(cfg)
    except Exception as e:
        return {"ok": False, "message": _friendly_error(e)}


def export_csv(cfg: dict[str, Any], sql: str) -> Iterator[bytes]:
    if cfg["engine"] == "mysql":
        return _mysql_export_csv(cfg, sql)
    return _ch_export_csv(cfg, sql)


def export_xlsx(cfg: dict[str, Any], sql: str, path: str) -> dict[str, Any]:
    if cfg["engine"] == "mysql":
        return _mysql_export_xlsx(cfg, sql, path)
    return _ch_export_xlsx(cfg, sql, path)


def _friendly_error(e: Exception) -> str:
    msg = str(e).strip()
    if "idna" in msg.lower() or "label too long" in msg.lower():
        return (
            "Invalid hostname or IP address. If you are using a Google Cloud SQL "
            "Instance Connection Name (project:region:instance), please connect via its IP "
            "or run the Cloud SQL Auth Proxy locally."
        )
    return msg or e.__class__.__name__
